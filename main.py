import streamlit as st
import pandas as pd
import unicodedata
from io import BytesIO

st.set_page_config(page_title="Separador de Planilha (qualf/nr)", layout="wide")

def normalize(s: str) -> str:
    s = unicodedata.normalize("NFD", str(s))
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    return " ".join(s.lower().strip().split())

# --- writer com autoajuste + congelar + autofiltro ---
def to_excel_bytes(dfx: pd.DataFrame, sheet_name: str) -> BytesIO:
    out = BytesIO()
    # escolhe o writer disponível
    try:
        import xlsxwriter  # noqa
        engine = "xlsxwriter"
    except Exception:
        try:
            import openpyxl  # noqa
            engine = "openpyxl"
        except Exception:
            raise RuntimeError("Instale: pip install xlsxwriter OU pip install openpyxl")

    with pd.ExcelWriter(out, engine=engine) as writer:
        dfx.to_excel(writer, index=False, sheet_name=sheet_name)

        if engine == "xlsxwriter":
            ws = writer.sheets[sheet_name]
            ws.freeze_panes(1, 0)
            ws.autofilter(0, 0, dfx.shape[0], dfx.shape[1]-1)
            for i, col in enumerate(dfx.columns):
                maxlen = max(len(str(col)), (dfx[col].astype(str).map(len).max() if not dfx.empty else 0))
                ws.set_column(i, i, min(maxlen + 2, 80))
        else:
            from openpyxl.utils import get_column_letter
            ws = writer.sheets[sheet_name]
            ws.freeze_panes = "A2"
            last_col = get_column_letter(dfx.shape[1])
            last_row = dfx.shape[0] + 1
            ws.auto_filter.ref = f"A1:{last_col}{last_row}"
            for i, col in enumerate(dfx.columns, start=1):
                maxlen = max(len(str(col)), (dfx[col].astype(str).map(len).max() if not dfx.empty else 0))
                ws.column_dimensions[get_column_letter(i)].width = min(maxlen + 2, 80)

    out.seek(0)
    return out

# --- escolha do engine para leitura ---
def pick_engine(filename: str) -> str:
    name = (filename or "").lower()
    if name.endswith(".xlsx"):
        return "openpyxl"   # precisa de openpyxl
    if name.endswith(".xls"):
        return "xlrd"       # precisa de xlrd==1.2.0
    return "openpyxl"

st.title("Separar Excel em QUALF e NR")

file = st.file_uploader("Envie o Excel (.xlsx ou .xls)", type=["xlsx", "xls"])

if file:
    # escolhe engine e tenta abrir a pasta de trabalho
    engine = pick_engine(file.name)
    try:
        xls = pd.ExcelFile(file, engine=engine)
    except ImportError as e:
        # mensagem amigável no deploy se faltar dependência
        missing = "openpyxl>=3.1.4" if engine == "openpyxl" else "xlrd==1.2.0"
        st.error(f"Dependência ausente para ler o arquivo ({engine}). Adicione no requirements: {missing}. Detalhe: {e}")
        st.stop()

    sheet = st.selectbox("Escolha a aba da planilha", xls.sheet_names, index=0)
    df = pd.read_excel(xls, sheet_name=sheet, engine=engine)

    # Normalizações de nomes
    norm = {c: normalize(c) for c in df.columns}
    def has_inativo(c): return "inativo" in normalize(c)

    # Colunas fixas
    targets = {
        "cargo":"cargo",
        "identificação":"identificacao",
        "nome da unidade superior":"nome da unidade superior",
        "id da unidade superior":"id da unidade superior",
        "usuários":"usuarios",
        "negócio da posição":"negocio da posicao",
    }

    # normalizado -> original
    nm2orig = {}
    for c,n in norm.items():
        nm2orig.setdefault(n, c)

    base_cols = []
    for _, key in targets.items():
        if key in nm2orig:
            base_cols.append(nm2orig[key])
        else:
            cand = [c for c,n in norm.items() if key in n]
            if cand: base_cols.append(cand[0])

    # Remove colunas com "inativo"
    valid_cols = [c for c in df.columns if not has_inativo(c)]
    base_cols = [c for c in base_cols if c in valid_cols]

    # Grupos
    qualf_cols = [c for c in valid_cols if normalize(c).startswith("qualf")]
    nr_cols    = [c for c in valid_cols if normalize(c).startswith("nr")]

    # Conjuntos finais (preservando ordem e sem duplicatas)
    def unique(seq): return list(dict.fromkeys(seq))
    df_qualf = df[unique(base_cols + qualf_cols)]
    df_nr    = df[unique(base_cols + nr_cols)]

    st.subheader("Prévia – QUALF")
    st.dataframe(df_qualf.head(20), use_container_width=True)
    st.subheader("Prévia – NR")
    st.dataframe(df_nr.head(20), use_container_width=True)

    st.download_button(
        "Baixar QUALF.xlsx",
        to_excel_bytes(df_qualf, "qualf"),
        file_name="qualf.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    st.download_button(
        "Baixar NR.xlsx",
        to_excel_bytes(df_nr, "nr"),
        file_name="nr.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
